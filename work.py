import openpyxl
import os
from datetime import datetime


class Work:

    def __init__(self, filename):
        self.filename = filename


    def addexpense(self,amt, name, filename):
        wb_obj = openpyxl.load_workbook(filename)
        sheet_obj = wb_obj.active
        for i in range(1, 100):
            values = sheet_obj.cell(row=i, column=3).value
            if values is None:
                x = i
                break
        cell_obj = sheet_obj.cell(row=x, column=2)
        cell_obj.value = amt
        cell_obj = sheet_obj.cell(row=x, column=3)
        cell_obj.value = name
        y = datetime.now()
        cell_obj = sheet_obj.cell(row=x, column=1)
        cell_obj.value = y
        wb_obj.save(filename)

    def addincome(self, amt, name, filename):
        wb_obj = openpyxl.load_workbook(filename)
        sheet_obj = wb_obj.active
        for i in range(1, 100):
            values = sheet_obj.cell(row=i, column=7).value
            if values is None:
                x = i
                break
        cell_obj = sheet_obj.cell(row=x, column=6)
        cell_obj.value = amt
        cell_obj = sheet_obj.cell(row=x, column=7)
        cell_obj.value = name
        y = datetime.now()
        cell_obj = sheet_obj.cell(row=x, column=5)
        cell_obj.value = y
        wb_obj.save(filename)

    def sumOfExpense(self, filename):
        wb_obj = openpyxl.load_workbook(filename)
        sheet_obj = wb_obj.active

        sum = 0
        for i in range(2, 100):
            y = sheet_obj.cell(row=i, column=2).value
            if y is None:
                break
            sum = sum + int(y)
        return sum

    def sumOfIncome(self, filename):
        wb_obj = openpyxl.load_workbook(filename)
        sheet_obj = wb_obj.active

        sum = 0
        for i in range(2, 100):
            y = sheet_obj.cell(row=i, column=6).value
            if y is None:
                break
            sum = sum + int(y)
        return sum

    def remaning(self, filename):
        wb_obj = openpyxl.load_workbook(filename)
        sheet_obj = wb_obj.active

        sum1 = 0
        for i in range(2, 100):
            y = sheet_obj.cell(row=i, column=2).value
            if y is None:
                break
            sum1 = sum1 + int(y)

        sum2 = 0
        for i in range(2, 100):
            y = sheet_obj.cell(row=i, column=6).value
            if y is None:
                break
            sum2 = sum2 + int(y)
        return sum2 - sum1

    def reset(self, filename):
        wb_obj = openpyxl.load_workbook(filename)
        sheet_obj = wb_obj.active

        for i in range(1, 9):
            for j in range(2, 100):
                cell_obj = sheet_obj.cell(row=j, column=i)
                if cell_obj is None:
                    continue
                else:
                    cell_obj.value = None
        wb_obj.save(filename)

    @staticmethod
    def create_worksheet(name):
        wb = openpyxl.Workbook()
        wb.title = name
        ws = wb.active
        ws.title = "Sheet1"
        wb.save(filename=name)

        wb_obj = openpyxl.load_workbook(name)
        sheet_obj = wb_obj.active

        cell_obj = sheet_obj.cell(row=1, column=1)
        cell_obj.value = "Date and Time"

        cell_obj = sheet_obj.cell(row=1, column=2)
        cell_obj.value = "Amount"

        cell_obj = sheet_obj.cell(row=1, column=3)
        cell_obj.value = "Application"

        cell_obj = sheet_obj.cell(row=1, column=5)
        cell_obj.value = "Date and Time"

        cell_obj = sheet_obj.cell(row=1, column=6)
        cell_obj.value = "Amount"

        cell_obj = sheet_obj.cell(row=1, column=7)
        cell_obj.value = "Source"
        wb_obj.save(name)
